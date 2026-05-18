"""MTO V1 — Excel export endpoint."""

import html
import io
import zipfile
import xml.etree.ElementTree as ET
import sqlite3

import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse

from dashboard.project_db import get_project_conn

try:
    import cairosvg as _cairosvg
    _HAS_CAIROSVG = True
except Exception:
    _HAS_CAIROSVG = False

router = APIRouter()

# Image dimensions used for embedding
_IMG_W = 480
_IMG_H = 360
_IMG_ANCHOR = "A3"
# Row offset for label EMU calculation: rows 1+2 at default 15pt height
_IMG_ROW_OFFSET_EMU = 2 * 15 * 12700  # 381000
_EMU_PER_PX = 9525
# Approximate rows the image occupies in the sheet (for table placement)
_IMG_ROWS = 22

_MAT_COLS = ["TAG", "part_description", "size", "material", "uom", "quantity", "total"]
_MAT_HEADERS = {
    "TAG": "TAG", "part_description": "Description", "size": "Size",
    "material": "Material", "uom": "UoM", "quantity": "Qty", "total": "Total",
}
_HDR_FILL = PatternFill("solid", fgColor="4472C4")
_HDR_FONT = Font(bold=True, color="FFFFFF")


# ── DB helpers ────────────────────────────────────────────────────────────────

def _get_typicals(conn: sqlite3.Connection, tool_id: int) -> list:
    return conn.execute(
        "SELECT * FROM mto_typicals WHERE tool_id = ? ORDER BY position, id", (tool_id,)
    ).fetchall()


def _get_materials(conn: sqlite3.Connection, typical_id: int) -> list:
    return conn.execute(
        "SELECT TAG, part_description, size, material, uom, quantity"
        " FROM mto_materials WHERE typical_id = ? ORDER BY position, id",
        (typical_id,),
    ).fetchall()


def _get_utility_count(conn: sqlite3.Connection, tool_id: int, typical_name: str) -> int:
    row = conn.execute(
        "SELECT COUNT(*) FROM mto_utilities WHERE tool_id = ? AND typical_name = ?",
        (tool_id, typical_name),
    ).fetchone()
    return row[0] if row else 0


def _get_image(conn: sqlite3.Connection, typical_id: int):
    return conn.execute(
        "SELECT filename, format, content FROM mto_images WHERE typical_id = ? LIMIT 1",
        (typical_id,),
    ).fetchone()


def _get_placements(conn: sqlite3.Connection, typical_id: int) -> list:
    return conn.execute(
        "SELECT tag, label_x, label_y FROM mto_tag_placements WHERE typical_id = ?",
        (typical_id,),
    ).fetchall()


# ── SVG conversion ────────────────────────────────────────────────────────────

def _svg_to_png(data: bytes) -> bytes | None:
    if not _HAS_CAIROSVG:
        return None
    try:
        return _cairosvg.svg2png(bytestring=data, output_width=_IMG_W, output_height=_IMG_H)
    except Exception:
        return None


# ── Sheet writer ──────────────────────────────────────────────────────────────

def _write_sheet(wb, conn: sqlite3.Connection, tool_id: int, typical) -> tuple[str, list]:
    """Add one worksheet for a typical. Returns (sheet_title, placements_for_injection)."""
    typical_id = typical["id"]
    typical_name = typical["name"]
    ws_title = typical_name[:31]
    ws = wb.create_sheet(title=ws_title)

    ws["A1"] = typical_name
    ws["A1"].font = Font(bold=True, size=14)
    if typical["description"]:
        ws["A2"] = typical["description"]
        ws["A2"].font = Font(italic=True, size=10)

    # Image area
    img_row = _get_image(conn, typical_id)
    has_image = False
    if img_row:
        fmt = img_row["format"].lower()
        content = bytes(img_row["content"])
        note = None
        png_data = None

        if fmt == "svg":
            png_data = _svg_to_png(content)
            if png_data is None:
                note = "SVG image present — install libcairo to embed (see README)"
        elif fmt == "pdf":
            note = "PDF image (not embeddable in Excel)"
        elif fmt == "dxf":
            note = "DXF image (not embeddable in Excel)"

        if png_data:
            xl_img = XlImage(io.BytesIO(png_data))
            xl_img.width = _IMG_W
            xl_img.height = _IMG_H
            ws.add_image(xl_img, _IMG_ANCHOR)
            has_image = True
        elif note:
            ws.cell(row=3, column=1, value=note).font = Font(italic=True, color="888888")

    # Materials table
    tbl_row = 3 + _IMG_ROWS
    for col_i, key in enumerate(_MAT_COLS, 1):
        c = ws.cell(row=tbl_row, column=col_i, value=_MAT_HEADERS[key])
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = Alignment(horizontal="center")

    util_count = _get_utility_count(conn, tool_id, typical_name)
    for r_i, mat in enumerate(_get_materials(conn, typical_id), 1):
        qty = mat["quantity"] or 0
        vals = [
            mat["TAG"], mat["part_description"], mat["size"],
            mat["material"], mat["uom"], qty, qty * util_count,
        ]
        for col_i, v in enumerate(vals, 1):
            ws.cell(row=tbl_row + r_i, column=col_i, value=v)

    # Return placements only when image is embedded (labels need an image to overlay)
    placements = _get_placements(conn, typical_id) if has_image else []
    return ws_title, placements


# ── Text-box XML injection ────────────────────────────────────────────────────

_DRAWING_CLOSE = "</xdr:wsDr>"


def _label_xml(shape_id: int, x_emu: int, y_emu: int, tag: str) -> str:
    """Return an <xdr:absoluteAnchor> XML snippet for a tag label text box."""
    w = int(_IMG_W * _EMU_PER_PX * 0.12)
    h = int(_IMG_H * _EMU_PER_PX * 0.06)
    t = html.escape(str(tag or ""), quote=False)
    return (
        f'<xdr:absoluteAnchor>'
        f'<xdr:pos x="{x_emu}" y="{y_emu}"/>'
        f'<xdr:ext cx="{w}" cy="{h}"/>'
        f'<xdr:sp macro="" textlink="">'
        f'<xdr:nvSpPr>'
        f'<xdr:cNvPr id="{shape_id}" name="TextBox {shape_id}"/>'
        f'<xdr:cNvSpPr txBox="1"/>'
        f'<xdr:nvPr/>'
        f'</xdr:nvSpPr>'
        f'<xdr:spPr>'
        f'<a:xfrm><a:off x="{x_emu}" y="{y_emu}"/><a:ext cx="{w}" cy="{h}"/></a:xfrm>'
        f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
        f'<a:solidFill><a:srgbClr val="FFFF99"/></a:solidFill>'
        f'<a:ln><a:solidFill><a:srgbClr val="888800"/></a:solidFill></a:ln>'
        f'</xdr:spPr>'
        f'<xdr:txBody>'
        f'<a:bodyPr wrap="square" inset="45720"/>'
        f'<a:lstStyle/>'
        f'<a:p><a:r><a:rPr lang="en-US" sz="700" b="1"/><a:t>{t}</a:t></a:r></a:p>'
        f'</xdr:txBody>'
        f'</xdr:sp>'
        f'<xdr:clientData/>'
        f'</xdr:absoluteAnchor>'
    )


def _map_sheets_to_drawings(zin: zipfile.ZipFile) -> dict[str, str]:
    """Return {sheet_name: drawing_xml_path} for sheets that reference a drawing."""
    _WB_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    _R_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    _PKG_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"

    wb_root = ET.fromstring(zin.read("xl/workbook.xml"))
    wb_rels_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))

    r_id_to_sheet = {
        rel.get("Id"): rel.get("Target")
        for rel in wb_rels_root.findall(f".//{_PKG_NS}Relationship")
        if "worksheet" in rel.get("Type", "")
    }

    result: dict[str, str] = {}
    for sheet_el in wb_root.findall(f".//{_WB_NS}sheet"):
        name = sheet_el.get("name")
        r_id = sheet_el.get(f"{_R_NS}id")
        target = r_id_to_sheet.get(r_id, "")
        if not target:
            continue
        sheet_filename = target.split("/")[-1]
        rels_path = f"xl/worksheets/_rels/{sheet_filename}.rels"
        if rels_path not in zin.namelist():
            continue
        for rel in ET.fromstring(zin.read(rels_path)).findall(f".//{_PKG_NS}Relationship"):
            if "drawing" in rel.get("Type", "").lower():
                raw = rel.get("Target", "")
                if raw.startswith("../"):
                    result[name] = "xl/" + raw[3:]
                else:
                    result[name] = f"xl/worksheets/{raw}"
                break
    return result


def _inject_labels(xlsx_bytes: bytes, sheet_placements: dict[str, list]) -> bytes:
    """Inject tag label text boxes into each affected drawing XML via zipfile surgery."""
    buf = io.BytesIO(xlsx_bytes)
    out = io.BytesIO()
    with zipfile.ZipFile(buf, "r") as zin:
        sheet_to_drawing = _map_sheets_to_drawings(zin)
        modified: dict[str, bytes] = {}
        for sheet_name, placements in sheet_placements.items():
            drw_path = sheet_to_drawing.get(sheet_name)
            if not drw_path or drw_path not in zin.namelist():
                continue
            xml = (modified.get(drw_path) or zin.read(drw_path)).decode("utf-8")
            if _DRAWING_CLOSE not in xml:
                continue
            base_id = xml.count("<xdr:sp") + 1
            snippets = []
            for i, pl in enumerate(placements):
                x_emu = int(pl["label_x"] * _IMG_W * _EMU_PER_PX)
                y_emu = int(_IMG_ROW_OFFSET_EMU + pl["label_y"] * _IMG_H * _EMU_PER_PX)
                snippets.append(_label_xml(base_id + i, x_emu, y_emu, pl["tag"]))
            xml = xml.replace(_DRAWING_CLOSE, "".join(snippets) + _DRAWING_CLOSE, 1)
            modified[drw_path] = xml.encode("utf-8")

        with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                zout.writestr(item, modified.get(item.filename) or zin.read(item.filename))

    return out.getvalue()


# ── Build + send ──────────────────────────────────────────────────────────────

def _build_response(conn: sqlite3.Connection, tool_id: int, typicals: list, filename: str) -> StreamingResponse:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if not typicals:
        ws = wb.create_sheet("No Data")
        ws["A1"] = "No typicals found for this tool."

    sheet_placements: dict[str, list] = {}
    for typical in typicals:
        title, placements = _write_sheet(wb, conn, tool_id, typical)
        if placements:
            sheet_placements[title] = list(placements)

    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    if sheet_placements:
        xlsx_bytes = _inject_labels(xlsx_bytes, sheet_placements)

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("/{tool_id}/export/excel")
def export_excel_all(
    tool_id: int,
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    """Export all typicals for an MTO tool to a multi-sheet Excel workbook."""
    if not conn.execute(
        "SELECT 1 FROM _tools WHERE id = ? AND tool_type = 'mto'", (tool_id,)
    ).fetchone():
        raise HTTPException(status_code=404, detail="MTO tool not found")
    typicals = _get_typicals(conn, tool_id)
    return _build_response(conn, tool_id, typicals, "mto_export.xlsx")


@router.get("/{tool_id}/export/excel/{typical_id}")
def export_excel_one(
    tool_id: int,
    typical_id: int,
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    """Export a single typical to an Excel workbook."""
    typical = conn.execute(
        "SELECT * FROM mto_typicals WHERE id = ? AND tool_id = ?", (typical_id, tool_id)
    ).fetchone()
    if not typical:
        raise HTTPException(status_code=404, detail="Typical not found")
    name = typical["name"].replace("/", "_")
    return _build_response(conn, tool_id, [typical], f"mto_{name}.xlsx")
