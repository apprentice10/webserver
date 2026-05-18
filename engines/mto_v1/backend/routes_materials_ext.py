"""MTO V1 — grid-api v1 Extended contract for mto_materials."""
import io
import json
import sqlite3
from typing import Any, List, Optional
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from dashboard.project_db import get_project_conn, get_current_revision
from dashboard.utils import format_log_entry, append_log
from .routes_materials import (
    _require_typical, _require_row, _utility_count, _serialize_row, _EDITABLE
)

router = APIRouter()


# ---------------------------------------------------------------------------
# Rows — Insert above/below, copy-insert
# ---------------------------------------------------------------------------

class _InsertBody(BaseModel):
    placement: str  # "above" | "below"


@router.post("/{tool_id}/materials/{typical_id}/rows/{row_id}/insert")
def insert_row(
    tool_id: int,
    typical_id: int,
    row_id: int,
    body: _InsertBody,
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    typical = _require_typical(tool_id, typical_id, conn)
    _require_row(typical_id, row_id, conn)
    rev = get_current_revision(conn)

    rows = conn.execute(
        "SELECT id FROM mto_materials WHERE typical_id = ? ORDER BY position, id",
        (typical_id,),
    ).fetchall()
    ids = [r["id"] for r in rows]
    anchor_idx = ids.index(row_id)
    insert_at = anchor_idx + 1 if body.placement == "below" else anchor_idx

    cur = conn.execute(
        "INSERT INTO mto_materials (typical_id, tag, rev, position) VALUES (?, '', ?, ?)",
        (typical_id, rev, 0),
    )
    new_id = cur.lastrowid
    conn.execute("UPDATE mto_materials SET tag = ? WHERE id = ?", (str(new_id), new_id))

    ids.insert(insert_at, new_id)
    for pos, rid in enumerate(ids):
        conn.execute(
            "UPDATE mto_materials SET position = ? WHERE id = ? AND typical_id = ?",
            (pos, rid, typical_id),
        )
    conn.commit()

    row = dict(conn.execute(
        "SELECT id, tag, rev, log, part_description, size, material, uom, quantity, position"
        " FROM mto_materials WHERE id = ?", (new_id,)
    ).fetchone())
    ucount = _utility_count(tool_id, typical["name"], conn)
    return _serialize_row(row, ucount)


@router.post("/{tool_id}/materials/{typical_id}/rows/{row_id}/copy-insert")
def copy_insert_row(
    tool_id: int,
    typical_id: int,
    row_id: int,
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    typical = _require_typical(tool_id, typical_id, conn)
    src = _require_row(typical_id, row_id, conn)
    rev = get_current_revision(conn)

    rows = conn.execute(
        "SELECT id FROM mto_materials WHERE typical_id = ? ORDER BY position, id",
        (typical_id,),
    ).fetchall()
    ids = [r["id"] for r in rows]
    insert_at = ids.index(row_id) + 1

    cur = conn.execute(
        "INSERT INTO mto_materials (typical_id, tag, rev, position,"
        " part_description, size, material, uom, quantity)"
        " VALUES (?, '', ?, ?, ?, ?, ?, ?, ?)",
        (typical_id, rev, 0,
         src.get("part_description", ""), src.get("size", ""),
         src.get("material", ""), src.get("uom", ""),
         src.get("quantity", 0)),
    )
    new_id = cur.lastrowid
    conn.execute("UPDATE mto_materials SET tag = ? WHERE id = ?", (str(new_id), new_id))

    ids.insert(insert_at, new_id)
    for pos, rid in enumerate(ids):
        conn.execute(
            "UPDATE mto_materials SET position = ? WHERE id = ? AND typical_id = ?",
            (pos, rid, typical_id),
        )
    conn.commit()

    row = dict(conn.execute(
        "SELECT id, tag, rev, log, part_description, size, material, uom, quantity, position"
        " FROM mto_materials WHERE id = ?", (new_id,)
    ).fetchone())
    ucount = _utility_count(tool_id, typical["name"], conn)
    return _serialize_row(row, ucount)


# ---------------------------------------------------------------------------
# Rows — Batch operations
# ---------------------------------------------------------------------------

class _BatchUpdateBody(BaseModel):
    cells: List[dict]


@router.post("/{tool_id}/materials/{typical_id}/rows/batch-update")
def batch_update(
    tool_id: int,
    typical_id: int,
    body: _BatchUpdateBody,
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    typical = _require_typical(tool_id, typical_id, conn)
    rev = get_current_revision(conn)
    ucount = _utility_count(tool_id, typical["name"], conn)

    updated = []
    for cell in body.cells:
        row_id = cell.get("row_id")
        slug = cell.get("col_slug")
        value = cell.get("value")
        if not row_id or slug not in _EDITABLE:
            continue
        existing = conn.execute(
            "SELECT * FROM mto_materials WHERE id = ? AND typical_id = ?",
            (row_id, typical_id),
        ).fetchone()
        if not existing:
            continue
        existing = dict(existing)
        log_entry = format_log_entry(rev, slug, existing.get(slug), value)
        new_log = append_log(existing.get("log") or "", log_entry)
        conn.execute(
            f"UPDATE mto_materials SET {slug} = ?, rev = ?, log = ? WHERE id = ?",
            (value, rev, new_log, row_id),
        )
        row = dict(conn.execute(
            "SELECT id, tag, rev, log, part_description, size, material, uom, quantity, position"
            " FROM mto_materials WHERE id = ?", (row_id,)
        ).fetchone())
        updated.append(_serialize_row(row, ucount))

    conn.commit()
    return updated


class _BatchOpBody(BaseModel):
    operation: str  # "soft_delete" | "hard_delete" | "restore" | "keep"
    row_ids: List[int]


@router.post("/{tool_id}/materials/{typical_id}/rows/batch-op")
def batch_op(
    tool_id: int,
    typical_id: int,
    body: _BatchOpBody,
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    _require_typical(tool_id, typical_id, conn)
    if body.operation in ("soft_delete", "hard_delete"):
        for rid in body.row_ids:
            conn.execute(
                "DELETE FROM mto_materials WHERE id = ? AND typical_id = ?",
                (rid, typical_id),
            )
        conn.commit()
    return {"ok": True}


@router.post("/{tool_id}/materials/{typical_id}/rows/batch-remove-override")
def batch_remove_override(
    tool_id: int,
    typical_id: int,
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    # MTO materials have no ETL override concept — no-op
    return {"ok": True}


class _PasteBody(BaseModel):
    rows: List[dict]


@router.post("/{tool_id}/materials/{typical_id}/rows/paste")
def paste_rows(
    tool_id: int,
    typical_id: int,
    body: _PasteBody,
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    typical = _require_typical(tool_id, typical_id, conn)
    rev = get_current_revision(conn)
    ucount = _utility_count(tool_id, typical["name"], conn)

    max_pos = conn.execute(
        "SELECT COALESCE(MAX(position), 0) FROM mto_materials WHERE typical_id = ?",
        (typical_id,),
    ).fetchone()[0]

    created = []
    for i, cells in enumerate(body.rows):
        cur = conn.execute(
            "INSERT INTO mto_materials (typical_id, tag, rev, position,"
            " part_description, size, material, uom, quantity)"
            " VALUES (?, '', ?, ?, ?, ?, ?, ?, ?)",
            (typical_id, rev, max_pos + i + 1,
             cells.get("part_description", ""), cells.get("size", ""),
             cells.get("material", ""), cells.get("uom", ""),
             cells.get("quantity", 0)),
        )
        new_id = cur.lastrowid
        conn.execute("UPDATE mto_materials SET tag = ? WHERE id = ?", (str(new_id), new_id))
        row = dict(conn.execute(
            "SELECT id, tag, rev, log, part_description, size, material, uom, quantity, position"
            " FROM mto_materials WHERE id = ?", (new_id,)
        ).fetchone())
        created.append(_serialize_row(row, ucount))

    conn.commit()
    return created


# ---------------------------------------------------------------------------
# Audit (parses log column)
# ---------------------------------------------------------------------------

@router.get("/{tool_id}/materials/{typical_id}/audit")
def get_audit(
    tool_id: int,
    typical_id: int,
    row_tag: Optional[str] = None,
    col_slug: Optional[str] = None,
    limit: int = 200,
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    _require_typical(tool_id, typical_id, conn)
    query = "SELECT id, tag, log FROM mto_materials WHERE typical_id = ?"
    params: list = [typical_id]
    if row_tag:
        query += " AND tag = ?"
        params.append(row_tag)
    rows = conn.execute(query, params).fetchall()

    entries = []
    for row in rows:
        log_text = row["log"] or ""
        for line in log_text.splitlines():
            line = line.strip()
            if not line:
                continue
            # Log format: "REV|col_slug|old_val|new_val|ts" (from format_log_entry)
            parts = line.split("|")
            if len(parts) < 4:
                continue
            entry_col = parts[1] if len(parts) > 1 else ""
            if col_slug and entry_col != col_slug:
                continue
            entries.append({
                "id":        len(entries) + 1,
                "row_tag":   row["tag"],
                "col_slug":  entry_col,
                "old_value": parts[2] if len(parts) > 2 else None,
                "new_value": parts[3] if len(parts) > 3 else None,
                "changed_at": parts[4] if len(parts) > 4 else "",
                "source":    "user",
            })
            if len(entries) >= limit:
                break
        if len(entries) >= limit:
            break

    return entries


# ---------------------------------------------------------------------------
# Find & Replace
# ---------------------------------------------------------------------------

class _FindReplaceBody(BaseModel):
    search: str
    replacement: str
    match_case: bool = False
    match_entire_cell: bool = False
    scope: Optional[List[dict]] = None


@router.post("/{tool_id}/materials/{typical_id}/find_replace")
def find_replace(
    tool_id: int,
    typical_id: int,
    body: _FindReplaceBody,
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    _require_typical(tool_id, typical_id, conn)
    rows = conn.execute(
        "SELECT id, tag, rev, log, part_description, size, material, uom, quantity"
        " FROM mto_materials WHERE typical_id = ?",
        (typical_id,),
    ).fetchall()
    rev = get_current_revision(conn)
    replaced = 0

    scope_set = None
    if body.scope:
        scope_set = {(c["row_id"], c["col_slug"]) for c in body.scope}

    search = body.search if body.match_case else body.search.lower()

    for row in rows:
        row = dict(row)
        row_id = row["id"]
        for slug in _EDITABLE:
            if scope_set and (row_id, slug) not in scope_set:
                continue
            val = str(row.get(slug) or "")
            cmp = val if body.match_case else val.lower()

            if body.match_entire_cell:
                match = cmp == search
            else:
                match = search in cmp

            if not match:
                continue

            if body.match_entire_cell:
                new_val = body.replacement
            else:
                if body.match_case:
                    new_val = val.replace(body.search, body.replacement)
                else:
                    import re
                    new_val = re.sub(re.escape(body.search), body.replacement, val, flags=re.IGNORECASE)

            log_entry = format_log_entry(rev, slug, val, new_val)
            new_log = append_log(row.get("log") or "", log_entry)
            conn.execute(
                f"UPDATE mto_materials SET {slug} = ?, rev = ?, log = ? WHERE id = ?",
                (new_val, rev, new_log, row_id),
            )
            row[slug] = new_val
            row["log"] = new_log
            replaced += 1

    conn.commit()
    return {"replaced": replaced}


@router.get("/{tool_id}/materials/{typical_id}/column_values/{col_slug}")
def column_values(
    tool_id: int,
    typical_id: int,
    col_slug: str,
    prefix: Optional[str] = None,
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    _require_typical(tool_id, typical_id, conn)
    if col_slug not in _EDITABLE:
        return []
    query = f"SELECT DISTINCT {col_slug} FROM mto_materials WHERE typical_id = ? AND {col_slug} != ''"
    params: list = [typical_id]
    if prefix:
        query += f" AND {col_slug} LIKE ?"
        params.append(f"{prefix}%")
    query += " ORDER BY 1 LIMIT 100"
    rows = conn.execute(query, params).fetchall()
    return [str(r[0]) for r in rows if r[0] is not None]


# ---------------------------------------------------------------------------
# Export — Excel
# ---------------------------------------------------------------------------

@router.get("/{tool_id}/materials/{typical_id}/export/excel")
def export_excel(
    tool_id: int,
    typical_id: int,
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    typical = _require_typical(tool_id, typical_id, conn)
    ucount = _utility_count(tool_id, typical["name"], conn)

    cols = conn.execute(
        "SELECT name, slug FROM mto_material_columns WHERE tool_id = ? ORDER BY position",
        (tool_id,),
    ).fetchall()
    rows = conn.execute(
        "SELECT id, tag, rev, log, part_description, size, material, uom, quantity, position"
        " FROM mto_materials WHERE typical_id = ? ORDER BY position, id",
        (typical_id,),
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = typical["name"][:31]

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)

    col_names = [c["name"] for c in cols]
    col_slugs = [c["slug"] for c in cols]

    for ci, name in enumerate(col_names, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, 2):
        row = dict(row)
        qty = row.get("quantity") or 0
        row["total"] = qty * ucount
        for ci, slug in enumerate(col_slugs, 1):
            ws.cell(row=ri, column=ci, value=row.get(slug, ""))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"materials_{typical['name']}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Sort / Filter state (per-typical)
# ---------------------------------------------------------------------------

@router.get("/{tool_id}/materials/{typical_id}/sort-filter-state")
def get_sort_filter_state(
    tool_id: int,
    typical_id: int,
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    _require_typical(tool_id, typical_id, conn)
    row = conn.execute(
        "SELECT state FROM mto_sf_state WHERE typical_id = ?", (typical_id,)
    ).fetchone()
    if not row:
        return {"sort": [], "filters": {}}
    try:
        return json.loads(row["state"])
    except Exception:
        return {"sort": [], "filters": {}}


class _SFStateBody(BaseModel):
    sort: Optional[list] = None
    filters: Optional[dict] = None


@router.patch("/{tool_id}/materials/{typical_id}/sort-filter-state")
def set_sort_filter_state(
    tool_id: int,
    typical_id: int,
    body: _SFStateBody,
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    _require_typical(tool_id, typical_id, conn)
    state = json.dumps({"sort": body.sort or [], "filters": body.filters or {}})
    conn.execute(
        "INSERT OR REPLACE INTO mto_sf_state (typical_id, state) VALUES (?, ?)",
        (typical_id, state),
    )
    conn.commit()
    return {"ok": True}
