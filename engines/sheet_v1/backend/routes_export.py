"""
engine/routes_export.py
------------------------
Export endpoints (Excel). Extracted from engine/routes.py (P3-004).
"""

import io
import sqlite3
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse

from dashboard.project_db import get_project_conn
from . import service

router = APIRouter(prefix="/api/engines", tags=["engine"])


@router.get("/{tool_id}/export/excel")
def export_excel(
    tool_id: int,
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    tool = service.get_engine(conn, tool_id)
    tool_slug = tool["slug"]
    tool_name = tool["name"]

    columns = [
        c for c in service.get_columns(conn, tool_id)
        if c["slug"] != "log"
    ]

    rows = conn.execute(
        f'SELECT * FROM "{tool_slug}" ORDER BY __position ASC'
    ).fetchall()
    rows = [dict(r) for r in rows]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tool_name[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D6A9F")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col["name"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col["slug"]))

    for col_idx, col in enumerate(columns, start=1):
        approx_width = max(len(col["name"]) + 2, 12)
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = approx_width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in tool_name)
    filename = f"{safe_name}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
