"""
engine/routes.py
-----------------
Endpoints HTTP del Table Engine — thin layer su service.py ed etl.py.
"""

import io
import json
import sqlite3
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, Any
from datetime import datetime

from engine.project_db import get_project_conn
from engine import service
from engine.catalog import TOOL_CATALOG

router = APIRouter(prefix="/api/tools", tags=["engine"])


# ============================================================
# SCHEMI PYDANTIC
# ============================================================

class ToolCreate(BaseModel):
    name:            str
    slug:            Optional[str] = None
    tool_type:       Optional[str] = None
    icon:            Optional[str] = "📄"
    template_id:     Optional[int] = None
    default_columns: Optional[list[dict]] = None
    etl_sql:         Optional[str] = None


class ToolSettingsUpdate(BaseModel):
    name:         Optional[str] = None
    rev:          Optional[str] = None
    current_rev:  Optional[str] = None   # alias per compatibilità
    note:         Optional[str] = None
    query_config: Optional[Any] = None
    icon:         Optional[str] = None


class ToolResponse(BaseModel):
    id:          int
    name:        str
    slug:        str
    tool_type:   Optional[str]
    current_rev: str
    note:        Optional[str]
    icon:        Optional[str]
    project_id:  Optional[int] = None   # non in _tools, lo inseriamo nel route

    class Config:
        from_attributes = True


class TemplateCreate(BaseModel):
    type_slug:   str
    name:        str
    description: Optional[str] = None
    etl_sql:     str
    project_id:  Optional[int] = None
    tool_id:     Optional[int] = None


class TemplateResponse(BaseModel):
    id:          int
    type_slug:   str
    name:        str
    description: Optional[str]
    etl_sql:     str
    created_at:  Optional[datetime]
    project_id:  Optional[int] = None
    tool_id:     Optional[int] = None

    class Config:
        from_attributes = True


class ColumnCreate(BaseModel):
    name:     str
    slug:     str
    col_type: Optional[str] = "text"
    width:    Optional[int] = 120
    position: Optional[int] = None


class ColumnUpdate(BaseModel):
    name:     Optional[str] = None
    width:    Optional[int] = None
    position: Optional[int] = None
    col_type: Optional[str] = None
    formula:  Optional[str] = None


class ColumnWidthUpdate(BaseModel):
    width: int

class ColumnReorder(BaseModel):
    order: list[int]   # IDs colonne utente nel nuovo ordine


class ColumnResponse(BaseModel):
    id:        int
    tool_id:   int
    name:      str
    slug:      str
    col_type:  str
    width:     int
    position:  int
    is_system: bool
    formula:   Optional[str] = None

    class Config:
        from_attributes = True


class RowCreate(BaseModel):
    cells: dict[str, Any]


class CellUpdate(BaseModel):
    slug:  str
    value: Optional[str] = None


class PasteData(BaseModel):
    rows: list[dict[str, Any]]


class SqlQuery(BaseModel):
    sql: str


class EtlQuery(BaseModel):
    sql:   str
    label: Optional[str] = None


class FlagCreate(BaseModel):
    name:  str
    color: str = "#888888"


class FlagUpdate(BaseModel):
    name:  Optional[str] = None
    color: Optional[str] = None


class CellFlagEntry(BaseModel):
    row_tag:  str
    col_slug: str = ""


class CellFlagToggleRequest(BaseModel):
    flag_id: int
    cells:   list[CellFlagEntry]


# ============================================================
# HELPER — adatta il dict _tools al formato ToolResponse
# ============================================================

def _tool_to_response(tool: dict, project_id: int = None) -> dict:
    config = {}
    if tool.get("query_config"):
        try:
            config = json.loads(tool["query_config"])
        except Exception:
            pass
    return {
        "id":          tool["id"],
        "name":        tool["name"],
        "slug":        tool["slug"],
        "tool_type":   tool.get("tool_type"),
        "current_rev": tool.get("rev", "A"),
        "note":        tool.get("note"),
        "icon":        tool.get("icon", "📄"),
        "project_id":  project_id,
        "is_stale":    bool(tool.get("is_stale", 0)),
        "has_etl":     bool(config.get("etl_sql", "").strip()),
    }


# ============================================================
# CATALOGO E TEMPLATE
# ============================================================

@router.get("/types")
def get_tool_types():
    return TOOL_CATALOG


@router.get("/templates", response_model=list[TemplateResponse])
def list_templates(
    type_slug:  Optional[str] = Query(None),
    project_id: Optional[int] = Query(None),
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    return service.get_templates(conn, type_slug)


@router.post("/templates", response_model=TemplateResponse)
def create_template(
    data: TemplateCreate,
    project_id: Optional[int] = Query(None),
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    return service.create_template(
        conn,
        type_slug=data.type_slug,
        name=data.name,
        etl_sql=data.etl_sql,
        description=data.description,
    )


@router.delete("/templates/{template_id}")
def delete_template(
    template_id: int,
    project_id: Optional[int] = Query(None),
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    service.delete_template(conn, template_id)
    return {"ok": True}


# ============================================================
# FLAGS (project-level, must come before /{tool_id} routes)
# ============================================================

@router.get("/flags")
def list_flags(
    project_id: Optional[int] = Query(None),
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    rows = conn.execute(
        "SELECT id, name, color, is_system FROM _flags ORDER BY is_system ASC, name ASC"
    ).fetchall()
    return [dict(r) for r in rows]


@router.post("/flags", status_code=201)
def create_flag(
    data: FlagCreate,
    project_id: Optional[int] = Query(None),
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    from fastapi import HTTPException as _HTTP
    try:
        cur = conn.execute(
            "INSERT INTO _flags (name, color, is_system) VALUES (?, ?, 0)",
            (data.name.strip(), data.color),
        )
        conn.commit()
        row = conn.execute(
            "SELECT id, name, color, is_system FROM _flags WHERE id = ?", (cur.lastrowid,)
        ).fetchone()
        return dict(row)
    except sqlite3.IntegrityError:
        raise _HTTP(status_code=409, detail=f"Flag '{data.name}' already exists")


@router.patch("/flags/{flag_id}")
def update_flag(
    flag_id: int,
    data: FlagUpdate,
    project_id: Optional[int] = Query(None),
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    from fastapi import HTTPException as _HTTP
    flag = conn.execute(
        "SELECT id, name, color, is_system FROM _flags WHERE id = ?", (flag_id,)
    ).fetchone()
    if not flag:
        raise _HTTP(status_code=404, detail="Flag not found")
    if flag["is_system"] and data.name is not None:
        raise _HTTP(status_code=400, detail="Cannot rename system flags")
    new_name  = data.name.strip() if data.name  is not None else flag["name"]
    new_color = data.color         if data.color is not None else flag["color"]
    try:
        conn.execute("UPDATE _flags SET name = ?, color = ? WHERE id = ?", (new_name, new_color, flag_id))
        conn.commit()
        row = conn.execute(
            "SELECT id, name, color, is_system FROM _flags WHERE id = ?", (flag_id,)
        ).fetchone()
        return dict(row)
    except sqlite3.IntegrityError:
        raise _HTTP(status_code=409, detail=f"Flag name '{new_name}' already exists")


@router.delete("/flags/{flag_id}", status_code=204)
def delete_flag(
    flag_id: int,
    project_id: Optional[int] = Query(None),
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    from fastapi import HTTPException as _HTTP
    flag = conn.execute(
        "SELECT id, is_system FROM _flags WHERE id = ?", (flag_id,)
    ).fetchone()
    if not flag:
        raise _HTTP(status_code=404, detail="Flag not found")
    if flag["is_system"]:
        raise _HTTP(status_code=400, detail="Cannot delete system flags")
    conn.execute("DELETE FROM _flags WHERE id = ?", (flag_id,))
    conn.commit()


@router.post("/{tool_id}/cell-flags/toggle", status_code=200)
def toggle_cell_flags(
    tool_id:    int,
    project_id: int = Query(...),
    data:       CellFlagToggleRequest = ...,
    conn:       sqlite3.Connection = Depends(get_project_conn),
):
    from fastapi import HTTPException as _HTTP
    tool      = service.get_tool(conn, tool_id)
    tool_slug = tool["slug"]

    flag = conn.execute(
        "SELECT id, is_system FROM _flags WHERE id = ?", (data.flag_id,)
    ).fetchone()
    if not flag:
        raise _HTTP(status_code=404, detail="Flag not found")
    if flag["is_system"]:
        raise _HTTP(status_code=400, detail="Cannot assign system flags manually")

    row_tags = [c.row_tag for c in data.cells]
    placeholders = ",".join("?" * len(row_tags))
    existing = conn.execute(
        f"""SELECT row_tag, col_slug FROM _cell_flags
            WHERE tool_slug = ? AND flag_id = ?
            AND row_tag IN ({placeholders})""",
        [tool_slug, data.flag_id] + row_tags,
    ).fetchall()

    existing_set = {(r["row_tag"], r["col_slug"]) for r in existing}
    all_have = all((c.row_tag, c.col_slug) in existing_set for c in data.cells)

    if all_have:
        for c in data.cells:
            conn.execute(
                "DELETE FROM _cell_flags WHERE tool_slug=? AND row_tag=? AND col_slug=? AND flag_id=?",
                (tool_slug, c.row_tag, c.col_slug, data.flag_id),
            )
        action = "removed"
    else:
        for c in data.cells:
            if (c.row_tag, c.col_slug) not in existing_set:
                conn.execute(
                    "INSERT OR IGNORE INTO _cell_flags (tool_slug, row_tag, col_slug, flag_id) VALUES (?,?,?,?)",
                    (tool_slug, c.row_tag, c.col_slug, data.flag_id),
                )
        action = "added"

    conn.commit()
    return {"action": action, "flag_id": data.flag_id, "cells_affected": len(data.cells)}


# ============================================================
# TOOL
# ============================================================

@router.get("/project/{project_id}")
def list_tools(
    project_id: int,
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    tools = service.get_tools_for_project(conn)
    return [_tool_to_response(t, project_id) for t in tools]


@router.post("/project/{project_id}")
def create_tool(
    project_id: int,
    data: ToolCreate,
    conn: sqlite3.Connection = Depends(get_project_conn),
):
    tool = service.create_tool(
        conn=conn,
        name=data.name,
        slug=data.slug,
        tool_type=data.tool_type,
        icon=data.icon,
        template_id=data.template_id,
        default_columns=data.default_columns,
        etl_sql=data.etl_sql
    )
    return _tool_to_response(tool, project_id)


@router.get("/{tool_id}")
def get_tool(
    tool_id: int,
    project_id: int = Query(...),
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    tool = service.get_tool(conn, tool_id)
    return _tool_to_response(tool, project_id)


@router.patch("/{tool_id}/settings")
def update_tool_settings(
    tool_id: int,
    project_id: int = Query(...),
    data: ToolSettingsUpdate = ...,
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    payload = data.model_dump(exclude_unset=True)
    # alias current_rev → rev
    if "current_rev" in payload and "rev" not in payload:
        payload["rev"] = payload.pop("current_rev")
    elif "current_rev" in payload:
        payload.pop("current_rev")

    tool = service.update_tool_settings(conn, tool_id, payload)
    return _tool_to_response(tool, project_id)


# ============================================================
# COLONNE
# ============================================================

@router.get("/{tool_id}/columns")
def list_columns(
    tool_id: int,
    project_id: int = Query(...),
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    return service.get_columns(conn, tool_id)


@router.post("/{tool_id}/columns")
def add_column(
    tool_id: int,
    project_id: int = Query(...),
    data: ColumnCreate = ...,
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    return service.add_column(
        conn=conn,
        tool_id=tool_id,
        name=data.name,
        slug=data.slug,
        col_type=data.col_type,
        width=data.width,
        position=data.position
    )


@router.put("/{tool_id}/columns/reorder")
def reorder_columns(
    tool_id: int,
    project_id: int = Query(...),
    data: ColumnReorder = ...,
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    return service.reorder_columns(conn, tool_id, data.order)


@router.patch("/{tool_id}/columns/{column_id}")
def update_column(
    tool_id: int,
    column_id: int,
    project_id: int = Query(...),
    data: ColumnUpdate = ...,
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    return service.update_column(
        conn, tool_id, column_id, data.model_dump(exclude_unset=True)
    )


@router.delete("/{tool_id}/columns/{column_id}")
def delete_column(
    tool_id: int,
    column_id: int,
    project_id: int = Query(...),
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    return service.delete_column(conn, tool_id, column_id)


@router.patch("/{tool_id}/columns/{column_id}/width")
def update_column_width(
    tool_id: int,
    column_id: int,
    data: ColumnWidthUpdate,
    project_id: int = Query(...),
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    return service.update_column_width(conn, tool_id, column_id, data.width)


# ============================================================
# RIGHE
# ============================================================

@router.get("/{tool_id}/rows")
def list_rows(
    tool_id: int,
    project_id: int = Query(...),
    include_deleted: bool = Query(False),
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    return service.get_rows(conn, tool_id, project_id, include_deleted)


@router.post("/{tool_id}/rows")
def create_row(
    tool_id: int,
    project_id: int = Query(...),
    data: RowCreate = ...,
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    return service.create_row(conn, tool_id, project_id, data.cells)


@router.post("/{tool_id}/rows/paste")
def paste_rows(
    tool_id: int,
    project_id: int = Query(...),
    data: PasteData = ...,
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    return service.paste_rows(conn, tool_id, project_id, data.rows)


@router.patch("/{tool_id}/rows/{row_id}/cell")
def update_cell(
    tool_id: int,
    row_id: int,
    project_id: int = Query(...),
    data: CellUpdate = ...,
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    return service.update_cell(
        conn, tool_id, row_id, project_id, data.slug, data.value
    )


@router.post("/{tool_id}/rows/{row_id}/delete")
def soft_delete_row(
    tool_id: int,
    row_id: int,
    project_id: int = Query(...),
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    return service.soft_delete_row(conn, tool_id, row_id, project_id)


@router.post("/{tool_id}/rows/{row_id}/restore")
def restore_row(
    tool_id: int,
    row_id: int,
    project_id: int = Query(...),
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    return service.restore_row(conn, tool_id, row_id, project_id)


@router.delete("/{tool_id}/rows/{row_id}/override")
def remove_override(
    tool_id: int,
    row_id: int,
    col: str = Query(...),
    project_id: int = Query(...),
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    return service.remove_override(conn, tool_id, row_id, col, project_id)


@router.post("/{tool_id}/rows/{row_id}/keep")
def keep_row(
    tool_id: int,
    row_id: int,
    project_id: int = Query(...),
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    from fastapi import HTTPException as _HTTP
    from engine.project_db import audit
    tool = service.get_tool(conn, tool_id)
    slug = tool["slug"]
    row = conn.execute(f'SELECT tag FROM "{slug}" WHERE __id = ?', (row_id,)).fetchone()
    if not row:
        raise _HTTP(status_code=404, detail="Row not found")
    tag = row["tag"]
    flag = conn.execute("SELECT id FROM _flags WHERE name = 'ETL: Eliminated'").fetchone()
    if flag:
        conn.execute(
            "DELETE FROM _cell_flags WHERE tool_slug=? AND row_tag=? AND col_slug='' AND flag_id=?",
            (slug, tag, flag["id"])
        )
        audit(conn, slug, "KEEP_ROW", row_tag=tag)
        conn.commit()
    return {"kept": True, "row_tag": tag}


@router.post("/{tool_id}/rows/{row_id}/hard-delete")
def hard_delete_row(
    tool_id: int,
    row_id: int,
    project_id: int = Query(...),
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    return service.hard_delete_row(conn, tool_id, row_id, project_id)


# ============================================================
# SQL EDITOR
# ============================================================

@router.post("/{tool_id}/sql")
def run_sql(
    tool_id: int,
    project_id: int = Query(...),
    data: SqlQuery = ...,
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    sql = data.sql.strip()
    forbidden = ["drop ", "alter ", "truncate ", "attach ", "detach "]
    sql_lower = sql.lower()
    for keyword in forbidden:
        if keyword in sql_lower:
            from fastapi import HTTPException
            raise HTTPException(
                status_code=403,
                detail=f"Operazione non permessa: '{keyword.strip()}'"
            )
    try:
        cur = conn.execute(sql)
        if cur.description:
            columns = [d[0] for d in cur.description]
            rows    = [dict(zip(columns, row)) for row in cur.fetchall()]
            return {"columns": columns, "rows": rows}
        conn.commit()
        return {"rowcount": cur.rowcount}
    except Exception as e:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail=str(e))


# ============================================================
# EXPORT
# ============================================================

@router.get("/{tool_id}/export/excel")
def export_excel(
    tool_id: int,
    project_id: int = Query(...),
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    tool = service.get_tool(conn, tool_id)
    tool_slug = tool["slug"]
    tool_name = tool["name"]

    # Colonne visibili: tutte tranne log e colonne interne __
    columns = [
        c for c in service.get_columns(conn, tool_id)
        if c["slug"] != "log"
    ]

    # Righe attive (non cancellate — stanno nella tabella principale)
    rows = conn.execute(
        f'SELECT * FROM "{tool_slug}" ORDER BY __position ASC'
    ).fetchall()
    rows = [dict(r) for r in rows]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tool_name[:31]  # Excel limita a 31 caratteri

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D6A9F")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col["name"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.freeze_panes = "A2"

    # Dati
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col["slug"]))

    # Larghezza colonne approssimata
    for col_idx, col in enumerate(columns, start=1):
        approx_width = max(len(col["name"]) + 2, 12)
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = approx_width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = "".join(c if c.isalnum() or c in "-_ " else "_" for c in tool_name)
    filename = f"{safe_name}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ============================================================
# ETL
# ============================================================

@router.post("/{tool_id}/etl/preview")
def etl_preview(
    tool_id: int,
    project_id: int = Query(...),
    data: EtlQuery = ...,
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    from engine.etl import etl_preview as _preview
    return _preview(conn, tool_id, data.sql)


@router.post("/{tool_id}/etl/apply")
def etl_apply(
    tool_id: int,
    project_id: int = Query(...),
    data: EtlQuery = ...,
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    from engine.etl import etl_apply as _apply
    return _apply(conn, tool_id, data.sql)


@router.post("/{tool_id}/etl/run")
def etl_run(
    tool_id: int,
    project_id: int = Query(...),
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    from engine.etl import etl_run_saved
    return etl_run_saved(conn, tool_id)


@router.post("/{tool_id}/etl/save")
def etl_save(
    tool_id: int,
    project_id: int = Query(...),
    data: EtlQuery = ...,
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    from engine.etl import save_etl_version
    return save_etl_version(conn, tool_id, data.sql, data.label)


@router.get("/{tool_id}/etl/config")
def etl_config(
    tool_id: int,
    project_id: int = Query(...),
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    from engine.etl import get_etl_config
    return get_etl_config(conn, tool_id)


@router.patch("/{tool_id}/etl/config")
def etl_save_draft(
    tool_id: int,
    project_id: int = Query(...),
    data: EtlQuery = ...,
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    from engine.etl import etl_save_draft as _save_draft
    return _save_draft(conn, tool_id, data.sql)


@router.get("/{tool_id}/etl/schema")
def etl_schema(
    tool_id: int,
    project_id: int = Query(...),
    conn: sqlite3.Connection = Depends(get_project_conn)
):
    from engine.etl import get_etl_schema
    return get_etl_schema(conn, tool_id)
